# -*- coding: utf-8 -*-
"""
续班多维分析系统（Streamlit版）
运行：streamlit run app.py
"""
import re
import io
import json
import pandas as pd
import numpy as np
import streamlit as st
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="续班多维分析系统", layout="wide")

REQUIRED_COLS = ["课程名称", "班级名称", "校区", "上课时间", "任课老师", "学号",
                  "姓名", "性别", "课程年级", "科目", "类型", "手机号码", "交费时间"]

DEFAULT_CONFIG = {
    "label1": "暑假报名名单", "label2": "秋季续费验证",
    "w_normal": 1.0, "w_g6": 1.3, "w_g6e": 1.3, "w_elite": 1.2,
    "nodes_str": "2026-06-01,2026-07-01,2026-08-01",
    "selected_types": None,
}

# ============================================================
# 工具函数
# ============================================================
def clean_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()


def extract_grade_code(text):
    if pd.isna(text):
        return ""
    text = str(text).strip()
    m = re.match(r"^(\d{2})", text)
    return m.group(1) if m else text


def extract_class_type(text):
    if pd.isna(text):
        return ""
    text = str(text).strip()
    m = re.search(r"[\u4e00-\u9fff]", text)
    return text[m.start():] if m else text


@st.cache_data(show_spinner=False)
def load_and_clean(file_bytes, label):
    try:
        raw = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    except Exception as e:
        return None, f"读取{label}失败：{e}"

    missing = [c for c in REQUIRED_COLS if c not in raw.columns]
    if missing:
        return None, f"{label}缺少必需列：{missing}；实际列：{list(raw.columns)}"

    df = raw[REQUIRED_COLS].copy()
    for col in ["姓名", "学号", "任课老师", "手机号码", "校区", "性别", "班级名称"]:
        df[col] = df[col].apply(clean_str)
    df["交费时间"] = df["交费时间"].apply(clean_str)
    df["交费日期"] = pd.to_datetime(df["交费时间"], errors="coerce")

    critical = ["姓名", "学号", "任课老师", "课程年级", "类型", "科目", "校区"]
    before = len(df)
    mask_missing = df[critical].apply(lambda c: c.map(lambda v: v == "" or pd.isna(v))).any(axis=1)
    dropped = int(mask_missing.sum())
    df = df[~mask_missing].reset_index(drop=True)

    df["课程年级码"] = df["课程年级"].apply(extract_grade_code)
    df["班型"] = df["类型"].apply(extract_class_type)

    msg = f"{label}清洗完成：原始{before}条，有效{len(df)}条" + (f"，剔除关键字段缺失{dropped}条" if dropped else "")
    return df, msg


def judge_renewal(df1, df2):
    name_map = {}
    for name, grp in df2.groupby("姓名"):
        name_map[name] = dict(zip(grp["学号"], grp["交费日期"]))

    results, pay_dates = [], []
    for _, row in df1.iterrows():
        name, sid = row["姓名"], row["学号"]
        if name not in name_map:
            results.append("未续班")
            pay_dates.append(pd.NaT)
        elif sid in name_map[name]:
            results.append("续班成功")
            pay_dates.append(name_map[name][sid])
        else:
            results.append("学号不匹配")
            pay_dates.append(pd.NaT)

    df1 = df1.copy()
    df1["续班结果"] = results
    df1["续班交费日期"] = pay_dates
    return df1


def get_category(row):
    if row["续班结果"] != "续班成功":
        return None
    g6 = "06" in str(row["课程年级码"])
    elite = "精品班" in str(row["班型"])
    if g6 and elite:
        return "六年级精品班"
    if g6:
        return "六年级非精品班"
    if elite:
        return "精品班"
    return "普通"


# ============================================================
# 五大维度分析函数
# ============================================================
def teacher_analysis(detail, weights):
    d = detail.copy()
    d["类别"] = d.apply(get_category, axis=1)
    d["权重"] = d["类别"].map(lambda c: weights.get(c, 0) if c else 0)
    g = d.groupby("任课老师").agg(总人数=("姓名", "count"), 加权成功=("权重", "sum")).reset_index()
    g["续班率"] = g["加权成功"] / g["总人数"]
    avg_rate = g["加权成功"].sum() / g["总人数"].sum() if g["总人数"].sum() else 0
    avg_cnt = g["总人数"].mean()

    fig = px.scatter(g, x="总人数", y="续班率", hover_name="任课老师", text="任课老师",
                      title="维度一 教师分析：数据1学生人数 vs 续班率（加权）")
    fig.update_traces(textposition="top center")
    fig.add_hline(y=avg_rate, line_dash="dash", line_color="red",
                  annotation_text=f"全校平均续班率 {avg_rate:.1%}")
    fig.add_vline(x=avg_cnt, line_dash="dash", line_color="green",
                  annotation_text=f"平均人数 {avg_cnt:.1f}")
    fig.update_yaxes(tickformat=".0%")
    return g, fig


def campus_analysis(detail):
    g = detail.groupby("校区").agg(
        总人数=("姓名", "count"),
        成功人数=("续班结果", lambda s: (s == "续班成功").sum())
    ).reset_index()
    g["续班率"] = g["成功人数"] / g["总人数"]

    fig = px.scatter(g, x="续班率", y="校区", size="总人数", color="校区",
                      size_max=60, title="维度二 校区分析：续班率 vs 学生总数（气泡大小）")
    fig.update_xaxes(tickformat=".0%")
    return g, fig


def grade_classtype_analysis(detail):
    rows = []
    for (grade, ctype), grp in detail.groupby(["课程年级", "班型"]):
        total = len(grp)
        success = (grp["续班结果"] == "续班成功").sum()
        classes = grp["班级名称"].nunique()
        rows.append({
            "年级": grade, "班型": ctype, "总人数": total,
            "续班率": success / total if total else 0,
            "开班数": classes,
            "班均人数": total / classes if classes else 0,
        })
    df = pd.DataFrame(rows)
    fig = px.scatter(df, x="班均人数", y="续班率", size="总人数", color="班型",
                      facet_col="年级", facet_col_wrap=4, size_max=40,
                      title="维度三 年级内嵌班型分析（分面：年级）")
    fig.update_yaxes(tickformat=".0%")
    return df, fig


def timeliness_analysis(detail, nodes):
    nodes = sorted(pd.to_datetime(nodes))
    total = len(detail)
    success = detail[detail["续班结果"] == "续班成功"].dropna(subset=["续班交费日期"]).sort_values("续班交费日期")

    bounds = [None] + list(nodes) + [None]
    labels, rates = [], []
    for i in range(len(nodes) + 1):
        lo, hi = bounds[i], bounds[i + 1]
        if lo is None:
            cnt = success[success["续班交费日期"] <= hi].shape[0]
            lbl = f"≤{hi.date()}"
        elif hi is None:
            cnt = success[success["续班交费日期"] > lo].shape[0]
            lbl = f">{lo.date()}"
        else:
            cnt = success[(success["续班交费日期"] > lo) & (success["续班交费日期"] <= hi)].shape[0]
            lbl = f"{lo.date()}~{hi.date()}"
        labels.append(lbl)
        rates.append(cnt / total if total else 0)

    bar_df = pd.DataFrame({"区间": labels, "转化率": rates})
    fig_bar = px.bar(bar_df, x="区间", y="转化率", text="转化率",
                      title="维度四 独立区间续班转化率")
    fig_bar.update_traces(texttemplate="%{y:.1%}")
    fig_bar.update_yaxes(tickformat=".0%")

    success = success.sort_values("续班交费日期").copy()
    success["累计人数"] = range(1, len(success) + 1)
    success["累计续班率"] = success["累计人数"] / total if total else 0
    fig_line = px.line(success, x="续班交费日期", y="累计续班率",
                        title="维度四 累计续班率增长趋势", markers=True)
    fig_line.update_yaxes(tickformat=".0%")
    for n in nodes:
        fig_line.add_vline(x=n, line_dash="dot", line_color="gray")

    return bar_df, fig_bar, fig_line


def gender_analysis(detail):
    g = detail.groupby("性别").agg(
        总人数=("姓名", "count"),
        成功人数=("续班结果", lambda s: (s == "续班成功").sum())
    ).reset_index()
    g["续班率"] = g["成功人数"] / g["总人数"]
    fig = px.bar(g, x="性别", y="续班率", color="性别", text="续班率",
                 title="维度五 性别续班率对比")
    fig.update_traces(texttemplate="%{y:.1%}")
    fig.update_yaxes(tickformat=".0%")
    return g, fig
def school_analysis(detail):
    total = len(detail)
    success = (detail["续班结果"] == "续班成功").sum()
    not_renew = (detail["续班结果"] == "未续班").sum()
    mismatch = (detail["续班结果"] == "学号不匹配").sum()
    rate = success / total if total else 0

    # 方式一：学号不匹配 归入 未续班（二分类）
    df_binary = pd.DataFrame({
        "结果": ["续班成功", "未续班（含学号不匹配）"],
        "人数": [success, not_renew + mismatch]
    })
    fig_binary = px.pie(df_binary, names="结果", values="人数", hole=0.5,
                         title="全校续班率（学号不匹配计入未续班）")
    fig_binary.update_traces(textinfo="label+percent")

    # 方式二：三类明细（续班成功 / 未续班 / 学号不匹配）
    df_detail = pd.DataFrame({
        "结果": ["续班成功", "未续班", "学号不匹配"],
        "人数": [success, not_renew, mismatch]
    })
    fig_detail = px.pie(df_detail, names="结果", values="人数", hole=0.5,
                         title="全校续班率明细（三类）")
    fig_detail.update_traces(textinfo="label+percent")

    return total, success, rate, df_binary, fig_binary, df_detail, fig_detail

# ============================================================
# 图表/数据 分开展示组件（默认图表，按钮切换到数据）
# ============================================================
def render_section(key, figs, df, chart_height=720):
    """
    key：本区块唯一标识（用于session_state隔离）
    figs：该维度下所有需要展示的图表（list[plotly.Figure]）
    df：该维度对应的数据表
    chart_height：单个图表高度（px），图表默认放大铺满宽度
    """
    state_key = f"view_mode_{key}"
    if state_key not in st.session_state:
        st.session_state[state_key] = "chart"  # 默认展示图表

    col_a, col_b, col_spacer = st.columns([1, 1, 4])
    with col_a:
        if st.button("📈 图表", key=f"btn_chart_{key}", use_container_width=True,
                      type="primary" if st.session_state[state_key] == "chart" else "secondary"):
            st.session_state[state_key] = "chart"
    with col_b:
        if st.button("📋 数据", key=f"btn_data_{key}", use_container_width=True,
                      type="primary" if st.session_state[state_key] == "data" else "secondary"):
            st.session_state[state_key] = "data"

    st.write("")  # 轻微留白

    if st.session_state[state_key] == "chart":
        # 只展示图表，不渲染数据表，避免图表被压缩变小
        for fig in figs:
            fig.update_layout(height=chart_height, margin=dict(l=40, r=40, t=60, b=40))
            st.plotly_chart(fig, use_container_width=True)
    else:
        # 只展示数据，不渲染图表
        table_height = min(700, 40 * (len(df) + 1) + 40)
        st.dataframe(df, use_container_width=True, height=table_height)


# ============================================================
# Excel 导出（续班明细 + 老师统计公式表）
# ============================================================
def autofit_columns(ws, df, extra_width=2):
    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df.iloc[:, i - 1].astype(str)])
        ws.column_dimensions[get_column_letter(i)].width = max_len + extra_width


def write_sheet(wb, sheet_name, df, percent_cols=None):
    ws = wb.create_sheet(sheet_name)
    percent_cols = percent_cols or []
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for j, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, col_name in enumerate(df.columns, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=i, column=j, value=val)
            if col_name in percent_cols:
                cell.number_format = "0.00%"
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws, df)
    return ws


def build_teacher_stats(detail_df):
    d = detail_df.copy()
    d["类别"] = d.apply(get_category, axis=1)
    rows = []
    for teacher, g in d.groupby("任课老师"):
        total = len(g)
        success = (g["续班结果"] == "续班成功").sum()
        mismatch = (g["续班结果"] == "学号不匹配").sum()
        rows.append({
            "老师姓名": teacher, "总人数": total, "续班成功人数": success, "学号不匹配人数": mismatch,
            "普通续班人数": (g["类别"] == "普通").sum(),
            "六年级非精品班续班人数": (g["类别"] == "六年级非精品班").sum(),
            "六年级精品班续班人数": (g["类别"] == "六年级精品班").sum(),
            "精品班续班人数": (g["类别"] == "精品班").sum(),
        })
    return pd.DataFrame(rows).sort_values("续班成功人数", ascending=False).reset_index(drop=True)


def write_teacher_stats_sheet(wb, stats_df, weights):
    cols = ["老师姓名", "总人数", "续班成功人数", "学号不匹配人数",
            "普通续班人数", "六年级非精品班续班人数", "六年级精品班续班人数", "精品班续班人数",
            "续班率", "加权后续班人数", "加权续班率"]
    df = stats_df.copy()
    for c in ["续班率", "加权后续班人数", "加权续班率"]:
        df[c] = None
    ws = write_sheet(wb, "老师统计", df[cols], percent_cols=["续班率", "加权续班率"])

    w_normal = weights["普通"]
    w_g6 = weights["六年级非精品班"]
    w_g6e = weights["六年级精品班"]
    w_elite = weights["精品班"]
    for r in range(2, len(df) + 2):
        ws[f"I{r}"] = f"=IF(B{r}=0,0,C{r}/B{r})"
        ws[f"J{r}"] = f"=E{r}*{w_normal}+F{r}*{w_g6}+G{r}*{w_g6e}+H{r}*{w_elite}"
        ws[f"J{r}"].number_format = "0.00"
        ws[f"K{r}"] = f"=IF(B{r}=0,0,J{r}/B{r})"


def export_excel(detail_df, weights):
    wb = Workbook()
    wb.remove(wb.active)
    detail_cols = ["姓名", "学号", "任课老师", "校区", "性别", "课程年级", "班型", "科目",
                   "手机号码", "交费时间", "续班结果"]
    write_sheet(wb, "续班明细", detail_df[detail_cols])
    stats_df = build_teacher_stats(detail_df)
    write_teacher_stats_sheet(wb, stats_df, weights)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# Streamlit 界面
# ============================================================
st.title("📊 续班多维分析系统")

with st.sidebar:
    st.header("⓪ 配置文件")
    cfg_file = st.file_uploader("导入配置文件（config.json）", type=["json"], key="cfg_uploader")
    if cfg_file is not None:
        try:
            loaded_cfg = json.load(cfg_file)
            st.session_state["cfg"] = {**DEFAULT_CONFIG, **loaded_cfg}
            st.success("配置已导入")
        except Exception as e:
            st.error(f"配置文件解析失败：{e}")
    cfg = st.session_state.get("cfg", DEFAULT_CONFIG)

    st.header("① 数据上传")
    label1 = st.text_input("数据1 含义", value=cfg["label1"])
    label2 = st.text_input("数据2 含义（续班验证）", value=cfg["label2"])
    file1 = st.file_uploader(f"上传数据1（{label1}）", type=["xlsx"], key="f1")
    file2 = st.file_uploader(f"上传数据2（{label2}）", type=["xlsx"], key="f2")

    st.header("② 加权系数（老师维度专用）")
    w_normal = st.number_input("普通班", value=cfg["w_normal"], step=0.1)
    w_g6 = st.number_input("六年级非精品班", value=cfg["w_g6"], step=0.1)
    w_g6e = st.number_input("六年级精品班", value=cfg["w_g6e"], step=0.1)
    w_elite = st.number_input("精品班（非六年级）", value=cfg["w_elite"], step=0.1)
    WEIGHTS = {"普通": w_normal, "六年级非精品班": w_g6, "六年级精品班": w_g6e, "精品班": w_elite}

    st.header("③ 交费时间分析节点")
    nodes_str = st.text_input("多个日期用逗号分隔", value=cfg["nodes_str"])

    run_btn = st.button("🚀 开始计算", type="primary", use_container_width=True)

    st.divider()
    current_cfg = {
        "label1": label1, "label2": label2,
        "w_normal": w_normal, "w_g6": w_g6, "w_g6e": w_g6e, "w_elite": w_elite,
        "nodes_str": nodes_str,
        "selected_types": st.session_state.get("selected_types_saved"),
    }
    st.download_button("💾 保存当前配置为 config.json",
                        data=json.dumps(current_cfg, ensure_ascii=False, indent=2).encode("utf-8"),
                        file_name="config.json", mime="application/json",
                        use_container_width=True)

if "df1_raw" not in st.session_state:
    st.session_state.df1_raw = None
    st.session_state.df2_raw = None

if file1 and file2:
    if run_btn or st.session_state.df1_raw is None:
        df1, msg1 = load_and_clean(file1.getvalue(), "数据1")
        df2, msg2 = load_and_clean(file2.getvalue(), "数据2")
        if df1 is None:
            st.error(msg1); st.stop()
        if df2 is None:
            st.error(msg2); st.stop()
        st.info(msg1); st.info(msg2)
        st.session_state.df1_raw = df1
        st.session_state.df2_raw = df2

if st.session_state.df1_raw is not None:
    df1 = st.session_state.df1_raw
    df2 = st.session_state.df2_raw

    st.subheader("④ 班型过滤")
    all_types = sorted(df1["班型"].unique())
    saved_types = cfg.get("selected_types")
    default_types = [t for t in saved_types if t in all_types] if saved_types else all_types
    selected_types = st.multiselect("勾选参与计算的班型（默认全选）", all_types, default=default_types)
    st.session_state["selected_types_saved"] = selected_types

    df1_filtered = df1[df1["班型"].isin(selected_types)].reset_index(drop=True)
    if df1_filtered.empty:
        st.warning("过滤后数据1为空，请至少勾选一个班型")
        st.stop()

    detail = judge_renewal(df1_filtered, df2)

    try:
        nodes = [n.strip() for n in nodes_str.split(",") if n.strip()]
        nodes = pd.to_datetime(nodes)
    except Exception as e:
        st.error(f"时间节点解析失败：{e}")
        st.stop()

    tabs = st.tabs(["教师分析", "校区分析", "年级×班型分析", "交费时效分析", "性别分析", "全校整体", "数据导出"])

    with tabs[0]:
        t_df, t_fig = teacher_analysis(detail, WEIGHTS)
        render_section("teacher", [t_fig], t_df)

    with tabs[1]:
        c_df, c_fig = campus_analysis(detail)
        render_section("campus", [c_fig], c_df)

    with tabs[2]:
        gc_df, gc_fig = grade_classtype_analysis(detail)
        render_section("grade_class", [gc_fig], gc_df, chart_height=780)

    with tabs[3]:
        bar_df, fig_bar, fig_line = timeliness_analysis(detail, nodes)
        render_section("timeliness", [fig_bar, fig_line], bar_df)

    with tabs[4]:
        g_df, g_fig = gender_analysis(detail)
        render_section("gender", [g_fig], g_df)
    with tabs[5]:  # 新增：全校整体
        total, success, rate, df_binary, fig_binary, df_detail, fig_detail = school_analysis(detail)
        col1, col2, col3 = st.columns(3)
        col1.metric("总人数", total)
        col2.metric("续班成功人数", success)
        col3.metric("续班率", f"{rate:.1%}")

        st.write("#### 方式一：学号不匹配计入未续班")
        render_section("school_binary", [fig_binary], df_binary)

        st.write("#### 方式二：三类明细")
    render_section("school_detail", [fig_detail], df_detail)
    with tabs[6]:
        st.write("### Excel 结果（续班明细 + 老师统计，含公式）")
        excel_buf = export_excel(detail, WEIGHTS)
        st.download_button("⬇️ 下载 Excel 结果", data=excel_buf,
                            file_name="续班统计结果.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.write("### 交互式 HTML 图表报告（所有维度合并）")
        html_parts = [
            "<html><head><meta charset='utf-8'><title>续班多维分析报告</title></head><body>",
            t_fig.to_html(full_html=False, include_plotlyjs="cdn"),
            c_fig.to_html(full_html=False, include_plotlyjs=False),
            gc_fig.to_html(full_html=False, include_plotlyjs=False),
            fig_bar.to_html(full_html=False, include_plotlyjs=False),
            fig_line.to_html(full_html=False, include_plotlyjs=False),
            g_fig.to_html(full_html=False, include_plotlyjs=False),
            "</body></html>",
        ]
        html_report = "\n".join(html_parts)
        st.download_button("⬇️ 下载交互式 HTML 报告", data=html_report.encode("utf-8"),
                            file_name="续班多维分析报告.html", mime="text/html")
else:
    st.info("请在左侧上传数据1与数据2后点击「开始计算」")